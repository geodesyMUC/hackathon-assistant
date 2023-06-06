import os
import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime
from tkcalendar import DateEntry


DATABASE_FILE = "hackathon_user_db.xlsx"
LABEL_COLOR = "#dadcee"


def check_date(date_string: str) -> bool:
    """Check if the input date is valid."""
    try:
        datetime.strptime(date_string, "%d.%m.%Y")
        return True
    except ValueError:
        return False


def add_to_database() -> None:
    """Add the textbox contents to the database."""
    id_value = id_entry.get().strip()
    last_name_value = last_name_entry.get().strip()
    first_name_value = first_name_entry.get().strip()
    start_date_value = start_date_entry.get_date().strftime("%d.%m.%Y")
    roll_off_date_value = roll_off_date_entry.get_date().strftime("%d.%m.%Y")

    if not id_value or not last_name_value or not first_name_value or \
            not start_date_value or not roll_off_date_value:
        messagebox.showerror("Error", "All fields must be filled.")
        return

    if not check_date(start_date_value) or not check_date(roll_off_date_value):
        messagebox.showerror("Error", "Invalid date format. Please use dd.mm.YYYY.")
        return

    # Create a new workbook if the file doesn't exist
    if not os.path.isfile(DATABASE_FILE):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "User Data"
        # Write header row
        header = ["ID", "Last Name", "First Name", "Start Date", "Roll-off Date"]
        sheet.append(header)
    else:
        # Open the existing workbook
        workbook = load_workbook(DATABASE_FILE)
        sheet = workbook.active

    # Find the first empty row
    row = sheet.max_row + 1

    # Write the textbox contents to the Excel file
    sheet.cell(row=row, column=1).value = id_value
    sheet.cell(row=row, column=2).value = last_name_value
    sheet.cell(row=row, column=3).value = first_name_value
    sheet.cell(row=row, column=4).value = start_date_value
    sheet.cell(row=row, column=5).value = roll_off_date_value

    # Save the workbook
    workbook.save(DATABASE_FILE)
    messagebox.showinfo("Success", "Data added to the database.")

    # Clear the textboxes
    id_entry.delete(0, tk.END)
    last_name_entry.delete(0, tk.END)
    first_name_entry.delete(0, tk.END)
    start_date_entry.delete(0, tk.END)
    roll_off_date_entry.delete(0, tk.END)


# Create the main window
root = tk.Tk()
root.title("Data Entry App")
root.configure(bg=LABEL_COLOR)

# Create labels and entry fields
label_style = {"bg": LABEL_COLOR, "fg": "#000000"}

id_label = tk.Label(root, text="ID", **label_style)
id_label.grid(row=0, column=0, padx=5, pady=5)
id_entry = ttk.Entry(root)
id_entry.grid(row=0, column=1, padx=5, pady=5)

last_name_label = tk.Label(root, text="Last Name", **label_style)
last_name_label.grid(row=1, column=0, padx=5, pady=5)
last_name_entry = ttk.Entry(root)
last_name_entry.grid(row=1, column=1, padx=5, pady=5)

first_name_label = tk.Label(root, text="First Name", **label_style)
first_name_label.grid(row=2, column=0, padx=5, pady=5)
first_name_entry = ttk.Entry(root)
first_name_entry.grid(row=2, column=1, padx=5, pady=5)

start_date_label = tk.Label(root, text="Start Date", **label_style)
start_date_label.grid(row=3, column=0, padx=5, pady=5)
start_date_entry = DateEntry(root, date_pattern="dd.mm.yyyy")
start_date_entry.grid(row=3, column=1, padx=5, pady=5)

roll_off_date_label = tk.Label(root, text="Roll-off Date", **label_style)
roll_off_date_label.grid(row=4, column=0, padx=5, pady=5)
roll_off_date_entry = DateEntry(root, date_pattern="dd.mm.yyyy")
roll_off_date_entry.grid(row=4, column=1, padx=5, pady=5)

# Create the "Add to Database" button
add_button = ttk.Button(root, text="Add to Database", command=add_to_database)
add_button.grid(row=5, column=0, columnspan=2, padx=5, pady=5)

# Run the GUI
root.mainloop()
